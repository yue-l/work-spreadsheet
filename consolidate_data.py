import sys
import re
import openpyxl
from openpyxl.styles import PatternFill
import DataSanitizer

# regex for email
regex = '[a-zA-Z0-9.-]+@[a-zA-Z0-9]+.[a-zA-Z0-9]+'

def isValidEmail(cellValue):
    print(cellValue)
    if cellValue == None:
        return False
    elif re.search(regex, cellValue):
        return True
    return False

def find_specific_cell(lookUpValue):
    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCDEFGHIJKL":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == lookUpValue:
                print("Specific cell on position: {} has value: {}".format(cell_name, currentSheet[cell_name].value))
                return currentSheet[cell_name]
    return False

# Main Fix text function
def fix_cell_format(cellContent):
    # Fix leading and trailing whitespace
    cellContent = DataSanitizer.trim_extra_space(cellContent)
    return cellContent

# Main script starts

# load workbook
try:
    wb = openpyxl.load_workbook(sys.argv[1])
except:
    sys.exit("target file must be supplied..abort operation")

# get all sheet names
allSheetNames = wb.sheetnames

print("All sheet names {} " .format(allSheetNames))

# loop thru each sheet
for sheet in allSheetNames:
    print("\n\nCurrent sheet name is ******* {} \n" .format(sheet))
    currentSheet = wb[sheet]

    # Find email column
    emailCell = find_specific_cell("Email")
    print(emailCell)
    if (not emailCell):
        sys.exit("Spreadsheet must contain Email column")
    print("Email column is: {}".format(emailCell.column))

    maxRow = currentSheet.max_row
    maxCol = currentSheet.max_column
    for i in range(1,maxRow+1):
        for l in range(1,maxCol+1):
            if (isinstance(currentSheet.cell(row=i, column=l).value, str)):
                currentSheet.cell(row=i, column=l).value = fix_cell_format(currentSheet.cell(row=i, column=l).value)
                
            if (
                currentSheet.cell(row=i, column=l).column == emailCell.column and
                currentSheet.cell(row=i, column=l).value is not None and 
                type(currentSheet.cell(row=i, column=l).value) == str and
                i > 1
            ):
                print("test email: {}".format(currentSheet.cell(row=i, column=l).value))
                print(isValidEmail(currentSheet.cell(row=i, column=l).value))
                if (not isValidEmail(currentSheet.cell(row=i, column=l).value)):
                    currentSheet.cell(row=i, column=l).fill = PatternFill(start_color="ff1100", end_color="ff1100", fill_type = "solid")
            elif (
                currentSheet.cell(row=i, column=l).column == emailCell.column and
                type(currentSheet.cell(row=i, column=l).value) != str
            ):
                currentSheet.cell(row=i, column=l).fill = PatternFill(start_color="ff1100", end_color="ff1100", fill_type = "solid")

print("Consolidation completed..")
wb.save("out.xlsx")
